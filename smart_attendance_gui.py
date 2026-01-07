import cv2
import numpy as np
import face_recognition
import os
from datetime import datetime
import pandas as pd
import time
import matplotlib.pyplot as plt
import tkinter as tk
from tkinter import messagebox

# === Liveness Detection ===
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
eye_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_eye.xml")

# === Load Known Faces ===
path = 'faces'
images = []
student_info = []

for filename in os.listdir(path):
    if filename.endswith(('.jpg', '.jpeg', '.png')):
        img = cv2.imread(os.path.join(path, filename))
        if img is not None:
            images.append(img)
            base = os.path.splitext(filename)[0]
            if "_" in base:
                student_id, name = base.split("_", 1)
            else:
                student_id, name = "UnknownID", base
            student_info.append((student_id, name))

# Encode faces and keep valid info
def findEncodings(images, student_info):
    encodeList = []
    valid_info = []
    for img, info in zip(images, student_info):
        rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        enc = face_recognition.face_encodings(rgb)
        if enc:
            encodeList.append(enc[0])
            valid_info.append(info)
        else:
            print(f"  Face not detected for {info[1]} ({info[0]}). Skipped.")
    return encodeList, valid_info

encodeListKnown, student_info = findEncodings(images, student_info)
print(f" Encodings loaded for {len(encodeListKnown)} valid students out of {len(images)} images.")

# === Attendance Dashboard Function ===
def show_summary():
    file_name = "attendance.xlsx"
    
    if not os.path.exists(file_name):
        messagebox.showinfo("Attendance Dashboard", "No attendance records found yet.")
        return

    try:
        # Read all sheets and combine
        excel_data = pd.read_excel(file_name, sheet_name=None)
        df = pd.concat(excel_data.values(), ignore_index=True)
    except Exception as e:
        messagebox.showerror("Error", f"Unable to read attendance file:\n{e}")
        return

    if df.empty:
        messagebox.showinfo("Attendance Dashboard", "Attendance file is empty.")
        return

    # Standardize date format
    df["Date"] = pd.to_datetime(df["Date"]).dt.strftime('%Y-%m-%d')

    today = datetime.now().strftime('%Y-%m-%d')
    today_records = df[df["Date"] == today]
    students_present_today = today_records["Student ID"].nunique()

    summary_text = f"""
ATTENDANCE DASHBOARD
-------------------------------
Date: {today}
Students Present Today: {students_present_today}
-------------------------------
"""
    messagebox.showinfo("Attendance Dashboard", summary_text)

# === Excel Attendance Logger ===
def markAttendance(student_id, name):
    date_str = datetime.now().strftime('%Y-%m-%d')
    time_str = datetime.now().strftime('%H:%M:%S')
    file_name = "attendance.xlsx"

    new_entry = pd.DataFrame([[student_id, name, date_str, time_str]],
                             columns=["Student ID", "Name", "Date", "Time"])

    if os.path.exists(file_name):
        from openpyxl import load_workbook
        wb = load_workbook(file_name)
        sheet_names = wb.sheetnames

        with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            if date_str in sheet_names:
                existing_df = pd.read_excel(file_name, sheet_name=date_str, engine='openpyxl')
                if (existing_df['Student ID'] == student_id).any():
                    print(f" Already marked for {name} today")
                    return
                updated_df = pd.concat([existing_df, new_entry], ignore_index=True)
                updated_df.to_excel(writer, sheet_name=date_str, index=False)
            else:
                new_entry.to_excel(writer, sheet_name=date_str, index=False)
    else:
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            new_entry.to_excel(writer, sheet_name=date_str, index=False)

    print(f" Attendance marked for {name} ({student_id}) at {time_str}")

# === Main Attendance Function ===
def start_attendance():
    FACE_DISTANCE_THRESHOLD = 0.5
    previous_locations = []
    frame_count = 0
    correct_recognitions = 0
    false_accepts = 0
    false_rejects = 0
    unknown_count = 0
    detection_times = []
    marked_students = set()

    cap = cv2.VideoCapture(0)
    messagebox.showinfo("Smart Attendance", "Webcam started. Press 'Q' to stop.")

    while True:
        start_time = time.time()
        success, img = cap.read()
        if not success:
            break

        frame_count += 1
        img = cv2.flip(img, 1)
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        # Liveness detection
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)
        liveness_confirmed = False
        for (x, y, w, h) in faces:
            roi_gray = gray[y:y+h, x:x+w]
            eyes = eye_cascade.detectMultiScale(roi_gray)
            if len(eyes) >= 1:
                liveness_confirmed = True
                break

        # Face recognition
        imgS = cv2.resize(img, (0, 0), fx=0.25, fy=0.25)
        rgbS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
        facesCurFrame = face_recognition.face_locations(rgbS)
        encodesCurFrame = face_recognition.face_encodings(rgbS, facesCurFrame)

        recognized_in_frame = False

        for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
            if not encodeListKnown:
                continue

            faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
            matchIndex = np.argmin(faceDis)
            student_id, student_name = student_info[matchIndex]

            motion_detected = False
            if previous_locations:
                prev = previous_locations[-1]
                diff = np.linalg.norm(np.array(faceLoc) - np.array(prev))
                if diff > 3:
                    motion_detected = True
            previous_locations.append(faceLoc)

            if faceDis[matchIndex] < FACE_DISTANCE_THRESHOLD and (liveness_confirmed or motion_detected):
                name_display = student_name.upper()
                correct_recognitions += 1
                recognized_in_frame = True

                if student_id not in marked_students:
                    markAttendance(student_id, student_name)
                    marked_students.add(student_id)
            else:
                name_display = "Unknown"
                unknown_count += 1
                if liveness_confirmed:
                    false_accepts += 1

            y1, x2, y2, x1 = [v * 4 for v in faceLoc]
            color = (0, 255, 0) if name_display != "Unknown" else (0, 0, 255)
            cv2.rectangle(img, (x1, y1), (x2, y2), color, 2)
            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), color, cv2.FILLED)
            cv2.putText(img, name_display, (x1 + 6, y2 - 6),
                        cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 2)

            if student_id in marked_students:
                cv2.putText(img, "Already Marked", (x1 + 6, y1 - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 255), 2)

        if not recognized_in_frame and liveness_confirmed:
            false_rejects += 1

        detection_times.append(time.time() - start_time)
        cv2.imshow('Smart Attendance System', img)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

    # === Performance Summary ===
    accuracy = (correct_recognitions / frame_count) * 100 if frame_count else 0
    far = (false_accepts / frame_count) * 100
    frr = (false_rejects / frame_count) * 100
    avg_time = np.mean(detection_times)

    summary = f"""
=== PERFORMANCE SUMMARY ===
Frames Processed: {frame_count}
Correct Recognitions: {correct_recognitions}
Unknown Detections: {unknown_count}
False Accepts (FAR): {far:.2f}%
False Rejects (FRR): {frr:.2f}%
Accuracy: {accuracy:.2f}%
Avg Detection Time: {avg_time:.3f} sec
"""
    print(summary)
    messagebox.showinfo("Performance Summary", summary)

    # === Graphs ===
    plt.figure(figsize=(12, 5))
    plt.subplot(1, 2, 1)
    metrics = ['Accuracy', 'FAR', 'FRR']
    values = [accuracy, far, frr]
    plt.bar(metrics, values, color=['green', 'red', 'orange'])
    plt.title("Face Recognition System Performance")
    plt.ylabel("Percentage (%)")

    plt.subplot(1, 2, 2)
    plt.plot(detection_times, color='blue')
    plt.title("Detection Time per Frame")
    plt.xlabel("Frame Number")
    plt.ylabel("Time (seconds)")
    plt.tight_layout()
    plt.show()

# === Tkinter GUI (Fullscreen Friendly) ===
root = tk.Tk()
root.title("Smart Attendance Tracker")
root.state('zoomed')  # Start in fullscreen-like mode
root.config(bg="#f0f4f7")

# Top Frame
top_frame = tk.Frame(root, bg="#003366", height=150)
top_frame.pack(fill="x")
tk.Label(top_frame, text="Smart Attendance Tracker", font=("Helvetica", 36, "bold"),
         bg="#003366", fg="white").pack(pady=40)
tk.Label(top_frame, text="Using Face Recognition & Liveness Detection",
         font=("Helvetica", 16, "italic"), bg="#003366", fg="white").pack()

# Middle Frame
middle_frame = tk.Frame(root, bg="#f0f4f7")
middle_frame.pack(fill="both", expand=True, pady=50)

tk.Button(middle_frame, text="Start Attendance", font=("Arial", 20, "bold"),
          bg="#28a745", fg="white", padx=50, pady=20, bd=0, relief="raised",
          activebackground="#218838", activeforeground="white",
          command=start_attendance).pack(pady=30)

tk.Label(middle_frame, text="Press 'Q' in webcam window to stop tracking",
         font=("Arial", 14), bg="#f0f4f7", fg="#555555").pack(pady=15)

tk.Button(middle_frame, text="View Attendance Dashboard", font=("Arial", 20, "bold"),
          bg="#007bff", fg="white", padx=45, pady=20, bd=0, relief="raised",
          activebackground="#0069d9", activeforeground="white",
          command=show_summary).pack(pady=30)

# Bottom Frame
bottom_frame = tk.Frame(root, bg="#e9ecef", height=60)
bottom_frame.pack(fill="x", side="bottom")
tk.Label(bottom_frame, text="© 2025 Smart Attendance Tracker",
         font=("Arial", 12), bg="#e9ecef", fg="#333333").pack(pady=15)

root.mainloop()